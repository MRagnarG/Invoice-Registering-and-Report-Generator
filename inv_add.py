import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from inv_tools import date_converter, comma_check
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill


excel_file = "invoices.xlsx"


def file_create():

    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        headers = [
            "Invoice number",
            "Appointment_date",
            "Payment date",
            "Patient/Dependent",
            "Payer SSN",
            "Dependent SSN",
            "Who Paid",
            "Amount ($)",
            "Payment Method",
            "Registering Date",
        ]
        ws.append(headers)

    return wb, ws


def get_inputs():
    print("📋 Fyll following Invoice Information:")
    invoice_nr = input("Invoice number: ")
    appointment_date = date_converter("Appointment_date: ")
    payment_date = date_converter("Payment date: ")
    patient = input("Patient/Dependent: ")
    payer_SSN = input("Payer SSN: ")
    dependent_SSN = input("Dependent SSN: ")
    payer = input("Who Paid: ")
    amount = comma_check()
    payment_method = input("Payment method: ")
    d_registro = datetime.today().strftime("%d/%m/%Y")
    return [
        invoice_nr,
        appointment_date,
        payment_date,
        patient,
        payer_SSN,
        dependent_SSN,
        payer,
        amount,
        payment_method,
        d_registro,
    ]


def format_spreadsheet(ws):

    # Freezes header
    ws.freeze_panes = "A2"

    # Bold text in header
    bold_font = Font(bold=True)

    # Grey background to header
    header_background = PatternFill(
        start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
    )

    # Lignt text centrally
    align_centrally = Alignment(horizontal="center", vertical="center")

    # Borders
    border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )

    # Column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    # Apply format styless to fylld places
    for row in ws.iter_rows(
        min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in row:
            if cell.value is not None:
                cell.alignment = align_centrally
                cell.border = border

                if cell.row == 1:
                    cell.font = bold_font
                    cell.fill = header_background

    # Format column "Amount ($)" como moeda (4ª column)
    col_valor_index = 8
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        cell = row[col_valor_index - 1]
        if isinstance(cell.value, float):
            cell.number_format = "$ #,##0.00"

    # Higher height to header
    ws.row_dimensions[1].height = 25


def run():

    while True:
        wb, ws = file_create()
        dados = get_inputs()
        ws.append(dados)
        format_spreadsheet(ws)
        wb.save(excel_file)
        print("✅ Input registering successfull!")

        while True:
            resposta = (
                input("Want to register another invoice? (y/n): ")
                .strip()
                .lower()
            )
            if resposta in ["y", "n"]:
                break
            else:
                print(
                    "⚠️ Invalid answer. Write only 'y' for yes or 'n' for no."
                )
        if resposta == "n":
            print("Shutting down the program, see you later! 👋")
            break


if __name__ == "__main__":
    run()
