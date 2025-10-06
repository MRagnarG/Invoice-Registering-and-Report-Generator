import os
from datetime import datetime
from openpyxl import load_workbook
from collections import Counter, defaultdict
from typing import List

"""
generate_reports.py

Script to generate financial and statistical reports based on the 
'invoices.xlsx' spreadsheet.
Available reports:
- General monthly
- Per patient (monthly and yearly)
- General yearly
- Totals per patient
- Custom date range

Developed by Matheus.
"""

MONTHS = {
    1: "January",
    2: "February",
    3: "March",
    4: "April",
    5: "May",
    6: "June",
    7: "July",
    8: "August",
    9: "September",
    10: "October",
    11: "November",
    12: "December",
}


def show_menu():
    print("---------------------------------------")
    print("===== INVOICE REPORTS GENERATOR =====")
    print("---------------------------------------\n")
    print("1. General report for current month")
    print("2. Patient report (monthly/yearly)")
    print("3. General yearly report")
    print("4. Totals per patient in the year")
    print("5. Custom report by date range")
    print("6. Exit")


def load_data(file_path="invoices.xlsx"):
    if not os.path.exists(file_path):
        print("⚠️ Invoices file not found.")
        return []

    wb = load_workbook(file_path)
    ws = wb.active

    data = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            entry = {
                "invoice_number": row[0],
                "appointment_date": str(row[1]),
                "payment_date": str(row[2]),
                "patient/dependent": row[3],
                "payer_SSN": row[4],
                "dependent_SSN": row[5],
                "who_paid": row[6],
                "amount": float(row[7]),
                "payment_method": row[8],
                "registering_date": str(row[9]),
            }
            data.append(entry)

        except Exception as e:
            print(f"Error loading row: {row}\n{e}")
            continue

    return data


def save_and_print(file_name: str, lines: List[str]) -> None:
    """
    Save the lines into the specified file and print them on console.
    :param file_name: path (or name) of the output file
    :param lines: list of strings, each will be one line in the file and
    console
    """
    try:
        with open(file_name, "w", encoding="utf-8") as f:
            for line in lines:
                f.write(f"{line}\n")
        print(f"✅ Report saved at: {file_name}")
    except IOError as e:
        print(f"❌ Error saving report '{file_name}': {e}")
        return

    # Console print
    for line in lines:
        print(line)


def monthly_general_report(data, month, year):
    lines = []
    lines.append(f"Monthly Report - {month:02d}/{year}")

    total_value = 0
    total_invoices = 0
    payments = Counter()
    patients = Counter()

    for invoice in data:
        payment_date = datetime.strptime(invoice["payment_date"], "%d/%m/%Y")
        if payment_date.month == month and payment_date.year == year:
            total_invoices += 1
            total_value += invoice["amount"]
            payments[invoice.get("payment_method", "N/A")] += 1
            patients[invoice.get("patient/dependent", "Unknown")] += 1

    if total_invoices == 0:
        print(f"📆 No appointments recorded in {month:02d}/{year}.")
        return

    avg_value = total_value / total_invoices

    lines.append(f"===== 📊 MONTHLY REPORT - {month:02d}/{year} =====")
    lines.append(f"Total invoices issued: {total_invoices}")
    lines.append(f"Total received: $ {total_value:.2f}")
    lines.append(f"Average per appointment: $ {avg_value:.2f}")
    lines.append("\nPayment methods used:")

    for method, count in payments.items():
        lines.append(f"- {method}: {count}x")

    top_patient = patients.most_common(1)[0]  # (name, qty)
    lines.append(
        f"\n👤 Most attended patient: {top_patient[0]}, "
        + f"({top_patient[1]} appointments)"
    )

    file_name = f"monthly_report_{month:02d}_{year}.txt"
    save_and_print(file_name, lines)

    print("======================================\n")


def yearly_general_report(data, year):
    total_value = 0
    total_invoices = 0
    values_per_month = defaultdict(float)
    payments = Counter()
    lines = []

    for invoice in data:
        try:
            payment_date = datetime.strptime(
                invoice["payment_date"], " +%d/%m/%Y"
            )
        except Exception:
            continue

        if payment_date.year == year:
            value = invoice.get("amount", 0.0)
            month = payment_date.month
            values_per_month[month] += value
            payments[invoice.get("payment_method", "N/A")] += 1
            total_value += value
            total_invoices += 1

    if total_invoices == 0:
        print(f"📅 No payments recorded in year {year}.")
        return

    monthly_avg = total_value / 12
    highest_month = max(values_per_month, key=values_per_month.get)
    highest_value = values_per_month[highest_month]
    lowest_month = min(values_per_month, key=values_per_month.get)
    lowest_value = values_per_month[lowest_month]
    most_used_payment = payments.most_common(1)[0][0]

    # Report lines
    lines.append(f"===== 📊 GENERAL YEARLY REPORT - {year} =====")
    lines.append(f"Total invoices: {total_invoices}")
    lines.append(f"Total received: $ {total_value:.2f}")
    lines.append(f"Monthly average: $ {monthly_avg:.2f}")
    lines.append(
        f"Highest revenue month: {MONTHS[highest_month]} -"
        + f" $ {highest_value:.2f}"
    )
    lines.append(
        f"Lowest revenue month: {MONTHS[lowest_month]} - "
        + f"$ {lowest_value:.2f}"
    )
    lines.append(f"Most used payment method: {most_used_payment}")
    lines.append("")
    lines.append("Values per month:")
    for month in range(1, 13):
        lines.append(f"- {MONTHS[month]}: $ {values_per_month[month]:.2f}")

    # Save file
    file_name = f"yearly_general_report_{year}.txt"
    save_and_print(file_name, lines)


def patient_monthly_report(data, patient, month, year):
    total_value = 0
    total_invoices = 0
    transactions = []
    lines = []

    for invoice in data:
        payment_date = datetime.strptime(invoice["payment_date"], "%d/%m/%Y")

        if payment_date.month == month and payment_date.year == year:
            if invoice.get("patient/dependent", "").lower() == (
                patient.lower()
            ):
                n = {
                    "Invoice Number": invoice.get("invoice_number", "N/A"),
                    "Appointment Date": invoice.get("appointment_date", "N/A"),
                    "Payment Date": invoice.get("payment_date", "N/A"),
                    "Patient/Dependent": invoice.get(
                        "patient/dependent", "Unknown"
                    ),
                    "Payer CPF": invoice.get("payer_CPF", "N/A"),
                    "Dependent CPF": invoice.get("dependent_CPF", "N/A"),
                    "Amount": invoice.get("amount", 0.0),
                    "Who Paid": invoice.get("who_paid", "N/A"),
                    "Payment Method": invoice.get("payment_method", "N/A"),
                    "Record Date": invoice.get("record_date", "N/A"),
                }

                total_value += n["Amount"]
                total_invoices += 1
                transactions.append(n)

    if total_invoices == 0:
        print(f"📆 No payments from {patient} recorded in {month:02d}/{year}.")
        return

    lines.append(
        f"===== 📄 PATIENT REPORT: {patient.upper()} -"
        + f"{month:02d}/{year} ====="
    )
    lines.append(f"Total appointments: {total_invoices}")
    lines.append(f"Total paid in the month: $ {total_value:.2f}")
    lines.append("")

    for i, t in enumerate(transactions, start=1):
        lines.append(f"--- Appointment {i} ---")
        for key, value in t.items():
            if key == "Amount":
                lines.append(f"{key}: $ {value:.2f}")
            else:
                lines.append(f"{key}: {value}")
        lines.append("")

    file_name = (
        f"patient_report_{patient.lower().replace(' ', '_')}"
        + f"_{month:02d}_{year}.txt"
    )

    save_and_print(file_name, lines)

    return total_value, total_invoices, transactions


def patient_yearly_report(data, patient, year):
    total_value = 0
    total_invoices = 0
    transactions = []
    lines = []

    for invoice in data:
        # Correct the date format
        payment_date = datetime.strptime(invoice["payment_date"], "%d/%m/%Y")

        # Filter by year and patient (safe comparison)
        if (
            payment_date.year == year
            and invoice.get("patient/dependent", "").lower() == patient.lower()
        ):
            n = {
                "Invoice Number": invoice.get("invoice_number", "N/A"),
                "Appointment Date": invoice.get("appointment_date", "N/A"),
                "Payment Date": invoice.get("payment_date", "N/A"),
                "Patient/Dependent": invoice.get(
                    "patient/dependent", "Unknown"
                ),
                "Payer CPF": invoice.get("payer_CPF", "N/A"),
                "Dependent CPF": invoice.get("dependent_CPF", "N/A"),
                "Amount": invoice.get("amount", 0.0),
                "Who Paid": invoice.get("who_paid", "N/A"),
                "Payment Method": invoice.get("payment_method", "N/A"),
                "Record Date": invoice.get("record_date", "N/A"),
            }

            total_value += n["Amount"]
            total_invoices += 1
            transactions.append(n)

    if total_invoices == 0:
        print(f"📅 No payments from {patient} recorded in {year}.")
        return

    # Build the report content
    lines.append(
        "===== 📄 YEARLY PATIENT REPORT: "
        + f"{patient.upper()} - {year} ====="
    )
    lines.append(f"Total appointments in the year: {total_invoices}")
    lines.append(f"Total amount paid in the year: $ {total_value:.2f}")
    lines.append("")

    for i, t in enumerate(transactions, start=1):
        lines.append(f"--- Appointment {i} ---")
        for key, value in t.items():
            if key == "Amount":
                lines.append(f"{key}: $ {value:.2f}")
            else:
                lines.append(f"{key}: {value}")
        lines.append("")

    # Save as .txt
    file_name = (
        f"yearly_report_{patient.lower().replace(' ', '_')}" + f"_{year}.txt"
    )
    save_and_print(file_name, lines)

    # Return useful data in case it’s needed later
    return total_value, total_invoices, transactions


def totals_per_patient_report(data, year):
    totals_per_patient = defaultdict(float)
    overall_total = 0
    lines = []

    for invoice in data:
        payment_date = datetime.strptime(invoice["payment_date"], "%d/%m/%Y")
        if payment_date.year == year:
            patient = invoice.get("patient/dependent", "Unknown")
            value = invoice.get("amount", 0.0)
            totals_per_patient[patient] += value
            overall_total += value

    if not totals_per_patient:
        print(f"📅 No patient data recorded in {year}.")
        return

    # Patient who paid the most
    top_patient = max(totals_per_patient, key=totals_per_patient.get)
    top_value = totals_per_patient[top_patient]
    avg_per_patient = overall_total / len(totals_per_patient)

    # Report lines
    lines.append(f"===== 📋 TOTALS PER PATIENT REPORT - {year} =====")
    for name, total in sorted(
        totals_per_patient.items(), key=lambda x: x[1], reverse=True
    ):
        lines.append(f"{name}: $ {total:.2f}")

    lines.append("\nSummary:")
    lines.append(f"Overall total received: $ {overall_total:.2f}")
    lines.append(f"Total number of patients: {len(totals_per_patient)}")
    lines.append(f"Average spending per patient: $ {avg_per_patient:.2f}")
    lines.append(f"Top paying patient: {top_patient} ($ {top_value:.2f})")

    # Save report
    file_name = f"totals_per_patient_report_{year}.txt"
    save_and_print(file_name, lines)


def custom_period_report(data, start_date, end_date, patient=None):
    start_date = datetime.strptime(start_date, "%d/%m/%Y")
    end_date = datetime.strptime(end_date, "%d/%m/%Y")

    total_invoices = 0
    total_value = 0
    payments = Counter()
    patients = Counter()
    transactions = []
    lines = []

    for invoice in data:
        payment_date = datetime.strptime(invoice["payment_date"], "%d/%m/%Y")

        if start_date <= payment_date <= end_date:
            patient_name = invoice.get("patient/dependent", "Unknown")
            if patient and patient_name.lower() != patient.lower():
                continue  # ignore if filtering by patient and doesn’t match

            value = invoice.get("amount", 0.0)
            payments[invoice.get("payment_method", "N/A")] += 1
            patients[patient_name] += 1
            total_value += value
            total_invoices += 1

            transactions.append(
                {
                    "Invoice Number": invoice.get("invoice_number", "N/A"),
                    "Appointment Date": invoice.get("appointment_date", "N/A"),
                    "Payment Date": invoice.get("payment_date", "N/A"),
                    "Patient/Dependent": patient_name,
                    "Amount": value,
                    "Payment Method": invoice.get("payment_method", "N/A"),
                    "Who Paid": invoice.get("who_paid", "N/A"),
                    "Record Date": invoice.get("record_date", "N/A"),
                }
            )

    if total_invoices == 0:
        print("📅 No appointments found in the specified period.")
        return

    avg_value = total_value / total_invoices
    period_str = (
        f"{start_date.strftime('%Y/%m/%d')} to "
        + f"{end_date.strftime('%Y/%m/%d')}"
    )

    if patient:
        title = f"CUSTOM REPORT - {patient.upper()} - {period_str}"
        file_name = f"custom_report_{patient.lower().replace(' ', '_')}.txt"
    else:
        title = f"CUSTOM REPORT - ALL PATIENTS - {period_str}"
        file_name = "custom_report_general.txt"

    # Header
    lines.append(f"===== 📅 {title} =====")
    lines.append(f"Total appointments: {total_invoices}")
    lines.append(f"Total received: $ {total_value:.2f}")
    lines.append(f"Average per appointment: $ {avg_value:.2f}")

    # Payment methods
    lines.append("\nPayment methods used:")
    for method, count in payments.items():
        lines.append(f"- {method}: {count}x")

    # Most attended patient (only in general report)
    if not patient:
        top_patient = patients.most_common(1)[0]
        lines.append(
            f"\n👤 Most attended patient: {top_patient[0]}"
            + f"({top_patient[1]} appointments)"
        )

    # Appointment details
    for i, t in enumerate(transactions, start=1):
        lines.append(f"\n--- Appointment {i} ---")
        for key, value in t.items():
            if key == "Amount":
                lines.append(f"{key}: $ {value:.2f}")
            else:
                lines.append(f"{key}: {value}")

    # Save
    save_and_print(file_name, lines)
